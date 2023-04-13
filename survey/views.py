# survey/views.py

from io import BytesIO
import io
import json
import os
from django.conf import settings
from django.contrib.auth import login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.template.loader import get_template
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.shortcuts import render, redirect, reverse, get_object_or_404, render
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes, force_str
from django.contrib import messages
from datetime import datetime

from django.views import View
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Survey, Question, Choice, SurveyResponse
from .tokens import user_tokenizer
from .forms import RegistrationForm, SurveyResponseForm
from django.contrib.auth import get_user_model

import io
from django.http import HttpResponse
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

def generate_report(survey):
    response_data = []
    
    for question in survey.questions.all():
        if not question.istext:
            choices = question.choices.all()
            total_responses = question.question_responses.count()
            question_data = []
            
            for choice in choices:
                num_responses = choice.choices_selected.count()
                if total_responses == 0:
                    percentage = 0
                else:
                    percentage = num_responses / total_responses * 100
                
                question_data.append([choice.text, percentage])
                
            response_data.append([question.text, question_data])
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    
    i=0
    for question_data in response_data:
        question_text = question_data[0]
        choice_data = question_data[1]
        
        drawing = Drawing(400, 400)
        pie = Pie()
        pie.x = 100
        pie.y = 250-i
        pie.width = 150
        pie.height = 150
        
        data = []
        labels = []
        for choice in choice_data:
            data.append(choice[1])
            labels.append(choice[0])
            
        pie.data = data
        pie.labels = labels
        pie.slices.strokeWidth = 0.5
        pie.slices.popout = 5
        drawing.add(pie)

        # Create a custom legend with corresponding percentages
        legend_data = []
        for choice in choice_data:
            legend_data.append([choice[0], f"{choice[1]:.2f}%"])

        legend_table = Table(legend_data, colWidths=[100, 50])
        legend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ]))
        legend_table.wrapOn(c, 100, 400)
        legend_table.drawOn(c, 350, 550-(3*i))

        c.drawString(50, 700-(2.5*i), question_text)
        renderPDF.draw(drawing, c, 50, 250-(2*i))
        i+=120
        if i > 121:
            i=0
            c.showPage()
        
    c.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={survey.title}_Results.pdf'
    return response


User = get_user_model()

def update_surveys():
    now = datetime.now()
    expired_surveys = Survey.objects.filter(deadline__lt=now, expired=False)
    for survey in expired_surveys:
        user = survey.created_by
        points = survey.allocated_points
        survey.expired = True
        survey.allocated_points -= points
        if points > 0:
            user.points += points
        survey.save()
        user.save()

def home(request):
    return render(request, 'survey/home.html', {})

def SurveyListview(request):
    update_surveys()
    surveys = Survey.objects.filter(expired=False)
    user_responses = SurveyResponse.objects.filter(created_by=request.user)
    answered_surveys = [response.question.survey for response in user_responses]
    unanswered_surveys = surveys.exclude(
    id__in=[survey.id for survey in answered_surveys if survey]
    ).exclude(
        allocated_points=0
    )   
    sorted_surveys = sorted(unanswered_surveys, key=lambda survey: survey.priorityValue(), reverse=True)
    context = {'surveys': sorted_surveys}
    return render(request, 'survey/list_surveys.html', context)

def survey_submitted(request):
    return render(request, 'survey/survey_submitted.html')

def survey_expired(request):
    return render(request, 'survey/survey_expired.html')

def fill_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)
    if survey.expired == True:
        return HttpResponseRedirect(reverse('survey_expired'))
    questions = survey.questions.all()
    form = SurveyResponseForm(questions=questions)

    if request.method == 'POST':
        form = SurveyResponseForm(request.POST, questions=questions)
        if form.is_valid():
            for question, choice_id in form.cleaned_data.items():
                if question.startswith('question_'):
                    question_id = question.replace('question_', '')
                    question = Question.objects.get(id=question_id)
                    if question.istext:
                        choice = Choice.objects.create(text=choice_id, question_id=question_id)
                    else:    
                        choice = Choice.objects.get(id=choice_id)
                    if request.user.is_authenticated:
                        user = request.user
                        if request.POST.get('anonymous') == 'True':
                            SurveyResponse.objects.create(is_anonymous=True, question_id=question_id, choice=choice, created_by=user)
                        else:
                            SurveyResponse.objects.create(question_id=question_id, choice=choice, created_by=user)    
                    else:
                        SurveyResponse.objects.create(question_id=question_id, choice=choice)   
            if request.user.is_authenticated:             
                user.points += 1
                if request.POST.get("group") != "":
                    user.group_name = request.POST.get("group")
                user.save()
            survey.allocated_points -= 1    
            survey.save()
            return HttpResponseRedirect(reverse('survey_submitted'))

    context = {
        'survey': survey,
        'form': form,
    }
    return render(request, 'survey/fill_survey.html', context)

class SharePointsView(View):

    def post(self, request):
    
        context = {}
        username = request.POST.get('username')
        points = int(request.POST.get('points'))
        
        try:
            receiver = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.error(request, 'User does not exist')
            context['username_error'] = 'User does not exist'
            return render(request, 'survey/share_points.html', context)
        
        sender = request.user
        if points < 1 or points > sender.points:
            messages.error(request, 'Invalid points value')
            context['points_error'] = 'Invalid points value'
            return render(request, 'survey/share_points.html', context)
        
        sender.points -= points
        receiver.points += points
        sender.save()
        receiver.save()
        
        messages.success(request, f'Successfully shared {points} points with {receiver.username}')
        context['success'] = f'Successfully shared {points} points with {receiver.username}'
        return render(request, 'survey/profile.html', context)

    def get(self, request):

        context = {}
        return render(request, 'survey/share_points.html', context)

class RegisterView(View):
    def get(self, request):
        return render(request, 'survey/register.html', { 'form': RegistrationForm() })

    def post(self, request):
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_valid = False
            user.save()
            #token = user_tokenizer.make_token(user)
            #user_id = urlsafe_base64_encode(force_bytes(str(user.id)))
            #url = 'http://localhost:8000' + reverse('confirm_email', kwargs={'user_id': user_id, 'token': token})
            #message = get_template('survey/register_email.html').render({'confirm_url': url})
            #mail = EmailMessage('SurFill Email Confirmation', message, to=[user.email], from_email=settings.EMAIL_HOST_USER)
            #mail.content_subtype = 'html'
            #mail.send()

            return render(request, 'survey/login.html', {
              'form': AuthenticationForm(),
              'message': f'Welcome to SurFill {user.username}. You can now login'
            })

        return render(request, 'survey/register.html', { 'form': form })

# this view class was replaced by django.contrib.auth.views.LoginView
class LoginView(View):
    def get(self, request):
        return render(request, 'survey/login.html', { 'form':  AuthenticationForm() })

    def post(self, request):
        form = AuthenticationForm(request, data=request.POST)
        remember_me = form.cleaned_data['remember_me']  # get remember me data from cleaned_data of form
        if not remember_me:
            self.request.session.set_expiry(0)  # if remember me is 
            self.request.session.modified = True
        if form.is_valid():
            try:
                form.clean()
            except ValidationError:
                return render(
                    request,
                    'survey/login.html',
                    { 'form': form, 'invalid_creds': True }
                )

            login(request, form.get_user())

            return redirect(reverse('profile'))

        return render(request, 'survey/login.html', { 'form': form })


class ProfileView(LoginRequiredMixin, View):

    def get(self, request):
        surveys = Survey.objects.filter(created_by=request.user)

        context = {'surveys': surveys}
        update_surveys()
        return render(request, 'survey/profile.html', context)


class SurveyEditView(LoginRequiredMixin, View):
    
    def get(self, request, survey_id):
        survey = get_object_or_404(Survey, id=survey_id)
        context = {'survey': survey}
        return render(request, 'survey/edit_survey.html', context)

    def post(self, request, survey_id):

        oldsurvey = get_object_or_404(Survey, pk=survey_id)
        data = request.POST
        print(data.get('anonymous'))
        
        title = data.get('title')
        questions_json = data.getlist('questions')
        deadline = data.get('deadline')
        allocated_points = data.get('points')

        if allocated_points == '':
            allocated_points = 0

        valid = True
        context = {'survey': oldsurvey}

        if not title:
            valid = False
            context['title_error'] = 'title is required'

        if not deadline:
            valid = False
            context['deadline_error'] = 'deadline is required' 

        if not questions_json:
            valid = False
            context['questions_error'] = 'questions are required'
            
        if not valid:
            context['users'] = User.objects.all()
            return render(request, 'survey/edit_survey.html', context)

        if request.POST.get('anonymous') == 'True':
            survey = Survey.objects.create(responder_info_required= False, title=title, created_by=request.user, created_at = oldsurvey.created_at,
                                        deadline = deadline, allocated_points=(allocated_points + oldsurvey.allocated_points))
            oldsurvey.delete()

        else:
            survey = Survey.objects.create(title=title, created_by=request.user, created_at = oldsurvey.created_at,
                                       deadline = deadline, allocated_points=(int(allocated_points) + oldsurvey.allocated_points))
            oldsurvey.delete()                                      

        request.user.points -= int(allocated_points)
        request.user.save()    

        for question_json in questions_json:
            question_data = json.loads(question_json)
            question = Question.objects.create(text=question_data['text'], survey=survey)
            if not question_data['choices']:
                question.istext = True
                question.save()
            else:
                for choice_data in question_data['choices']:
                    Choice.objects.create(text=choice_data['text'], question=question)


        # Show success message and redirect to survey list
        messages.success(request, 'Survey updated successfully!')
        return redirect('profile')


class SurveyCreateView(LoginRequiredMixin, View):
    def get(self, request):
        users = User.objects.all()
        return render(request, 'survey/create_survey.html', {'users': users})
    
    def post(self, request):
        data = request.POST
        
        title = data.get('title')
        questions_json = data.getlist('questions')
        deadline = data.get('deadline')
        allocated_points = data.get('points')
        if allocated_points == '':
            allocated_points = 0
        valid = True
        context = {}
        if not title:
            valid = False
            context['title_error'] = 'title is required'

        if not deadline:
            valid = False
            context['deadline_error'] = 'deadline is required' 

        if not questions_json:
            valid = False
            context['questions_error'] = 'questions are required'
            
        if not valid:
            context['users'] = User.objects.all()
            return render(request, 'survey/create_survey.html', context)

        if request.POST.get('anonymous') == 'True':
            survey = Survey.objects.create(responder_info_required= False, title=title, created_by=request.user, 
                                        deadline = deadline, allocated_points=allocated_points)
        else:
            survey = Survey.objects.create(title=title, created_by=request.user, 
                                       deadline = deadline, allocated_points=allocated_points)

        request.user.points -= int(allocated_points)
        request.user.save()                               
        for question_json in questions_json:
            question_data = json.loads(question_json)
            question = Question.objects.create(text=question_data['text'], survey=survey)
            if not question_data['choices']:
                question.istext = True
                question.save()
            else:
                for choice_data in question_data['choices']:
                    Choice.objects.create(text=choice_data['text'], question=question)

        return redirect(reverse('profile'))


class QuestionViewModel:
    def __init__(self, text):
        self.text = text
        self.choices = []

    def add_survey_response(self, survey_response):
        for choice in self.choices:
            if choice.id == survey_response.choice.id:
                choice.responses += 1
                break


class ChoiceResultViewModel:
    def __init__(self, id, text, responses=0):
        self.id = id
        self.text = text
        self.responses = responses


class SurveyResultsView(View):

    def get_object(self):
        self.obj = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
        return self.obj

    def get(self, request, survey_id):
        self.obj = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
        questions = []
        for question in self.obj.questions.all():
            question_vm = QuestionViewModel(question.text)
            for choice in question.choices.all():
                question_vm.choices.append(ChoiceResultViewModel(choice.id, choice.text))
            
            for survey_response in SurveyResponse.objects.filter(question=question):
                question_vm.add_survey_response(survey_response)
            
            questions.append(question_vm)

        context = {'survey': self.obj, 'questions': questions}
        
        return render(request, 'survey/survey_results.html', context)

    def post(self, request, survey_id):
        if 'downloadpdf' in request.POST:
            self.obj = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
            return generate_report(self.obj)

        if 'download' in request.POST:
            self.obj = get_object_or_404(Survey, pk=self.kwargs['survey_id'])
            questions = []
            for question in self.obj.questions.all():
                question_vm = QuestionViewModel(question.text)
                for choice in question.choices.all():
                    question_vm.choices.append(ChoiceResultViewModel(choice.id, choice.text))
                
                for survey_response in SurveyResponse.objects.filter(question=question):
                    question_vm.add_survey_response(survey_response)
                
                questions.append(question_vm)

            for question in questions:
                print(question.text)
                for choice in question.choices:
                    print(choice.text, choice.responses)
                print()
            return export_survey_results(self.obj, questions)

        if 'delete' in request.POST:
            survey = self.get_object()
            if survey.allocated_points > 0:
                request.user.points += survey.allocated_points
                request.user.save()
            survey = Survey.objects.get(id=survey_id)
            survey.delete()
            return redirect('profile')
        survey = get_object_or_404(Survey, id=survey_id)
        points = int(request.POST.get('points', 0))
        if points > 0:
            survey.allocated_points += points
            survey.save()
            request.user.points -= int(points)
            request.user.save()
            return redirect('profile') 
        elif points != 0:
            survey.allocated_points -= abs(points)
            survey.save()
            request.user.points += abs(int(points))
            request.user.save()
            return redirect('profile')
        return redirect('survey_results', survey_id=survey_id)    


class TestEmail(View):
    def get(self, request):
        user = User.objects.get(pk=9)
        token = user_tokenizer.make_token(user)
        user_id = urlsafe_base64_encode(force_bytes(user.id))
        url = 'http://localhost:8000' + reverse('confirm_email', kwargs={'user_id': user_id, 'token': token})
        message = get_template('survey/register_email.html').render({
          'confirm_url': url
        })
        mail = EmailMessage('Django Survey Email Confirmation', message, to=[user.email], from_email=settings.EMAIL_HOST_USER)
        mail.content_subtype = 'html'
        mail.send()
        return HttpResponse(f'email sent user_id = {user_id}, token = {token}')


class ConfirmRegistrationView(View):
    def get(self, request, user_id, token):
        user_id = force_str(urlsafe_base64_decode(force_bytes(user_id)))
        
        user = User.objects.get(pk=user_id)

        context = {
          'form': AuthenticationForm(),
          'message': 'Registration confirmation error . Please click the reset password to generate a new confirmation email.'
        }
        if user and user_tokenizer.check_token(user, token):
            user.is_valid = True
            user.save()
            context['message'] = 'Registration complete. Please login'

        return render(request, 'survey/login.html', context)

def export_survey_results(survey, questions):
    wb = openpyxl.Workbook()
    
    # iterate over each question
    for question in survey.questions.all():
        # create a new sheet with the question text as the title
        sheet = wb.create_sheet(title=str(question.text).replace("?", ""))
        
        # add the table headers
        sheet['A1'] = 'Answer'
        sheet['B1'] = 'Email'
        sheet['C1'] = 'Username'
        sheet['D1'] = 'Group'
        
        
        # iterate over each response and add it to the sheet
        for index, response in enumerate(question.question_responses.all()):
            sheet.cell(row=index+2, column=1, value=response.choice.text)
            if (response.is_anonymous != True) and (response.created_by is not None):
                sheet.cell(row=index+2, column=2, value=(response.created_by.get_email()))
                sheet.cell(row=index+2, column=3, value=(response.created_by.get_username()))
                sheet.cell(row=index+2, column=4, value=(response.created_by.get_group()))
            else:
                sheet.cell(row=index+2, column=2, value=("Anonymous"))
                sheet.cell(row=index+2, column=3, value=("Anonymous"))
                sheet.cell(row=index+2, column=4, value=("Anonymous"))
            

        sheet.column_dimensions[get_column_letter(1)].width = 50
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        sheet.column_dimensions[get_column_letter(4)].width = 30
        
    
    first_sheet = wb.sheetnames[0]
    wb.remove(wb[first_sheet])

    # create a response object with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    title = survey.title + "_Results.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{title}"'
    wb.save(response)
    return response
